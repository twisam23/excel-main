import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
import os
from datetime import datetime
import pandas as pd
import json
import gspread
from gspread.exceptions import SpreadsheetNotFound
from oauth2client.service_account import ServiceAccountCredentials
from PIL import Image, ImageTk

class StartDialog:
    def __init__(self, root):
        self.root = root
        
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        self.root.geometry(f"{screen_width}x{screen_height}")

        self.create_start_dialog()

    def create_start_dialog(self):
        try:
            image = Image.open("C:/Users/vyoma/Downloads/excelapp-main (1)/excelapp-main/E STORE-logos.jpeg")
            image = image.resize((600, 600), Image.ANTIALIAS)
            logo_image = ImageTk.PhotoImage(image)

            window_width, window_height = image.size
            self.root.geometry(f"{window_width}x{window_height}")

            self.root.configure(background='#FFDAB9')
            self.root.overrideredirect(True)

            logo_label = ttk.Label(self.root, image=logo_image, background='#FFDAB9')
            logo_label.image = logo_image
            logo_label.pack()

            self.root.update_idletasks()
            width = self.root.winfo_width()
            height = self.root.winfo_height()
            x = (self.root.winfo_screenwidth() // 2) - (width // 2)
            y = (self.root.winfo_screenheight() // 2) - (height // 2)
            self.root.geometry(f"+{x}+{y}")

            self.root.after(3000, self.load_main_application)
        except Exception as e:
            print(f"Error: {e}")

    def center_window(self):
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x_coordinate = (screen_width - self.root.winfo_reqwidth()) / 2
        y_coordinate = (screen_height - self.root.winfo_reqheight()) / 2
        self.root.geometry("+%d+%d" % (x_coordinate, y_coordinate))

    def load_main_application(self):
        try:
            self.root.destroy()  # Close the loading screen
            root = tk.Tk()  # Create the main root window
            app = StoreApp(root)  # Initialize your main application window
            root.mainloop()  # Run the main loop
        except Exception as e:
            print("Error loading main application:", str(e))

class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip = None
        self.widget.bind("<Enter>", self.on_enter)
        self.widget.bind("<Leave>", self.on_leave)

    def show_tooltip(self):
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + self.widget.winfo_width() // 2
        y += self.widget.winfo_rooty() + self.widget.winfo_height() + 5

        self.tooltip = tk.Toplevel(self.widget)
        self.tooltip.wm_overrideredirect(True)
        self.tooltip.wm_geometry(f"+{x}+{y}")

        label = tk.Label(self.tooltip, text=self.text, background="#ffffe0", relief="solid", borderwidth=1)
        label.pack(ipadx=2)

    def on_enter(self, event=None):
        self.show_tooltip()

    def on_leave(self, event=None):
        if self.tooltip:
            self.tooltip.destroy()
            self.tooltip = None

class StoreApp:
    def __init__(self, root):
        self.root = root
        self.root.title('Store Supplies')
        
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        self.root.geometry(f"{screen_width}x{screen_height}")

        self.load_data_from_google_sheets()
        self.load_log_from_google_sheets()

        self.entry_added_in = None
        self.entry_person_name = None
        self.entry_used_for = None
        self.entry_person_name_2 = None

    def load_data_from_google_sheets(self):
        try:
            gc = self.authorize_google_sheets()
            if gc:
                sh = gc.open("testdata")  # Replace with your spreadsheet title
                worksheet = sh.sheet1  # Access a specific sheet if needed
                data = worksheet.get_all_values()
                self.elements = data[1:]  # Assuming the header is in the first row
                print(self.elements)
        except Exception as e:
            print("Error loading data file:", str(e))

    def load_log_from_google_sheets(self):
        try:
            gc = self.authorize_google_sheets()
            if gc:
                sh = gc.open("storingfile")  # Replace with your spreadsheet title
                worksheet = sh.sheet1  # Access a specific sheet if needed
                log_data = worksheet.get_all_values()
                self.log_data = pd.DataFrame(log_data[1:], columns=log_data[0])  # Assuming header in the first row
                print(self.log_data)
        except Exception as e:
            print("Error reading log file:", str(e))
    
    def authorize_google_sheets(self):
        try:
            scope = [
                'https://www.googleapis.com/auth/spreadsheets',
                'https://www.googleapis.com/auth/drive'
            ]
            creds = ServiceAccountCredentials.from_json_keyfile_name('excelapp-main\e-store-408305-2d24bf10c72e.json', scope)
            return gspread.authorize(creds)
        except Exception as e:
            print("Google Sheets authorization error:", str(e))
            
    def display_logs(self, storingfile_path):
        if storingfile_path and storingfile_path.endswith(('.xlsx', '.xls')):
            excel_data = pd.read_excel(storingfile_path)
            # Process excel_data as needed
            print(excel_data)
        else:
            print("Invalid file format or no file selected.")

    def search_selected_material(self, event=None):
        if self.combodata:
            value2 = self.combodata.get()

            if value2 == '':
                self.combodata['values'] = [item[0] for item in self.elements]
            else:
                data = []
                for item in self.elements:
                    if value2.lower() in item[0].lower():
                        data.append(item[0])

                self.combodata['values'] = data

    def load_stored_paths_from_spreadsheet(self, spreadsheet_path):
        try:
            # Assuming the file paths are stored in a specific sheet and columns (e.g., 'Paths', 'DataFilePath', 'LogFilePath')
            df = pd.read_excel(spreadsheet_path, sheet_name='Paths')

            # Retrieve the data and log file paths from the spreadsheet columns
            data_file_path = df['DataFilePath'].iloc[0]  # Assuming it's in the first row
            log_file_path = df['LogFilePath'].iloc[0]  # Assuming it's in the first row

            if data_file_path and log_file_path:
                # Check if the stored paths point to existing files
                if os.path.isfile(data_file_path) and os.path.isfile(log_file_path):
                    self.data_file_path = data_file_path
                    self.log_file_path = log_file_path
                    self.show_bf1_dialogue()
                    self.load_data_file()  # Initialize self.wb
                    self.read_log_file()  # Initialize self.log_data
                    return  # Added to exit function after showing BF-1 dialogue
                else:
                    print("Stored paths do not point to existing files.")

            # If paths are missing or invalid, show the initial dialogue
            if not (self.data_file_path and self.log_file_path):
                self.create_initial_dialogue()

        except FileNotFoundError:
            # If the spreadsheet file is not found or the sheet doesn't exist, show the initial dialogue
            self.create_initial_dialogue()

    def show_bf1_dialogue(self):
        self.root.withdraw()
        bf1_dialog = tk.Toplevel(self.root)
        bf1_dialog.title("Store Supplies")

        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        bf1_dialog.geometry(f"{screen_width}x{screen_height}")

        heading_label = ttk.Label(bf1_dialog, text="Store Supplies", font=('Arial', 16))
        heading_label.pack(pady=20)

        display_button = tk.Button(
            bf1_dialog,  # Change this line to use bf1_dialog instead of self.root
            text='Blast Furnace-1',
            command=self.show_bf1_options,
            font=('Arial', 16),
            bg='blue',
            fg='white',
            padx=20,
            pady=10,
            relief=tk.RAISED,
            borderwidth=5
        )
        display_button.place(relx=0.5, rely=0.45, anchor="center")


            #calling the tooltip/button description
        bf1tooltip_text = "Click to enter Blast Furnace-1 options"
        ToolTip(display_button, bf1tooltip_text)

            # Initialize combobox
        self.combodata = None

    
    def show_bf1_options(self):
        # Close the current dialogue box
        self.root.withdraw()

        # Create a new dialog for BF-1 options
        bf1_dialog = tk.Toplevel(self.root)
        bf1_dialog.title("Blast Furnace-1 Options")

        # Create and style the heading
        heading_label = ttk.Label(bf1_dialog, text="Choose Blast Furnace-1 Option", font=('Arial', 16))
        heading_label.pack(pady=20)

        #finding the max window dimensions
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        # Set the main window size to fill the screen
        bf1_dialog.geometry(f"{screen_width}x{screen_height}")

        self.load_data_file()

        # Create a button to Add materials
        add_button = tk.Button(
            bf1_dialog,
            text='Add Materials',
            command=self.add_materials_dialog,
            font=('Arial', 13),  # Change font and size
            fg='black',  # Foreground (text) color
            padx=20,  # Horizontal padding
            pady=10,  # Vertical padding
            relief=tk.RAISED,  # Border style
            borderwidth=5  # Border width
        )
        add_button.place(relx=0.5, rely=0.2, anchor="center")

        #calling the tooltip/button description
        addtooltip_text = "Click to add new Materials to the store"
        ToolTip(add_button, addtooltip_text)

        # Remove Materials Button
        remove_button = tk.Button(
            bf1_dialog,
            text='Remove Materials',
            command=self.remove_materials_dialog,
            font=('Arial', 13),  # Change font and size
            fg='black',  # Foreground (text) color
            padx=20,  # Horizontal padding
            pady=10,  # Vertical padding
            relief=tk.RAISED,  # Border style
            borderwidth=5  # Border width
        )
        remove_button.place(relx=0.5, rely=0.3, anchor="center")

        #calling the tooltip/button description
        removetooltip_text = "Click to take/remove materials from the store"
        ToolTip(remove_button, removetooltip_text)

        # Material Status Button
        status_button = tk.Button(
            bf1_dialog,
            text='Material Status',
            command=self.display_material_status,
            font=('Arial', 13),  # Change font and size
            fg='black',  # Foreground (text) color
            padx=20,  # Horizontal padding
            pady=10,  # Vertical padding
            relief=tk.RAISED,  # Border style
            borderwidth=5  # Border width
        )
        status_button.place(relx=0.5, rely=0.4, anchor="center")

        #calling the tooltip/button description
        statustooltip_text = "Click to view Store Material status"
        ToolTip(status_button, statustooltip_text)

        # Requirements logs Button
        logs_button = tk.Button(
            bf1_dialog,
            text='Store logs',
            command=self.display_logs,
            font=('Arial', 13),  # Change font and size
            fg='black',  # Foreground (text) color
            padx=20,  # Horizontal padding
            pady=10,  # Vertical padding
            relief=tk.RAISED,  # Border style
            borderwidth=5  # Border width
        )
        logs_button.place(relx=0.5, rely=0.5, anchor="center")

        #calling the tooltip/button description
        logstooltip_text = "Click to view Store logs"
        ToolTip(logs_button, logstooltip_text)


    def add_materials_dialog(self):
        # Create a new dialog for Add Materials
        add_dialog = tk.Toplevel(self.root)
        add_dialog.title("Add Materials")
        
        # Create and style the heading
        heading_label = ttk.Label(add_dialog, text="Choose materials to add to the store", font=('Arial', 16))
        heading_label.pack(pady=20)

        #finding the max window dimensions
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        # Set the main window size to fill the screen
        add_dialog.geometry(f"{screen_width//2}x{screen_height//2}")

        # Create a frame to hold label and Combobox for material
        frame = tk.Frame(add_dialog)
        frame.pack(padx=10, pady=10, anchor=tk.W)

        # Frame for Material Description
        frame_material_desc = tk.Frame(frame)
        frame_material_desc.pack(anchor=tk.W)

        # label and ComboBox for material selection
        self.material_label = ttk.Label(frame_material_desc, text="Material Description: ", font=('Arial', 10), width=18, anchor=tk.W)
        self.material_label.pack(side=tk.LEFT, padx=10, pady=10)
        self.combodata = ttk.Combobox(frame_material_desc, values=[item[0] for item in self.elements], width=50)
        self.combodata.pack(side=tk.LEFT, padx=10, pady=10)
        self.combodata.bind('<KeyRelease>', self.search_selected_material)
        #calling the tooltip/button description
        materialtooltip_text = "Select materials to add"
        ToolTip(self.combodata, materialtooltip_text)

        # Frame for Quantity
        frame_quantity = tk.Frame(frame)
        frame_quantity.pack(anchor=tk.W)

        # Entry for quantity input
        self.quantity_label = ttk.Label(frame_quantity, text="Quantity: ", font=('Arial', 10), width=18, anchor=tk.W)
        self.quantity_label.pack(side=tk.LEFT, padx=10, pady=10)
        entry_quantity = ttk.Entry(frame_quantity, width = 15)
        entry_quantity.pack(padx=10, pady=10)
        #calling the tooltip/button description
        qtytooltip_text = "Enter integer quantity"
        ToolTip(entry_quantity, qtytooltip_text)

        frame_entries = tk.Frame(add_dialog)
        frame_entries.pack(padx=10, pady=10, anchor=tk.W)

        self.added_in_label = ttk.Label(frame_entries, text="Added In: ", font=('Arial', 10), width=12, anchor=tk.E)
        self.added_in_label.grid(row=0, column=0, padx=10, pady=5)
        self.entry_added_in = ttk.Entry(frame_entries, width=20)
        self.entry_added_in.grid(row=0, column=1, padx=10, pady=5)

        self.person_name_label = ttk.Label(frame_entries, text="Person Name: ", font=('Arial', 10), width=12, anchor=tk.E)
        self.person_name_label.grid(row=1, column=0, padx=10, pady=10)
        self.entry_person_name = ttk.Entry(frame_entries, width=20)
        self.entry_person_name.grid(row=1, column=1, padx=10, pady=10)

        # Button to confirm adding materials
        confirm_button = ttk.Button(add_dialog, text="Confirm", command=lambda: self.handle_action("add", entry_quantity.get()))
        confirm_button.pack(pady=20, anchor=tk.CENTER)


    def remove_materials_dialog(self):
        # Create a new dialog for Remove Materials
        remove_dialog = tk.Toplevel(self.root)
        remove_dialog.title("Remove Materials")

        # Create and style the heading
        heading_label = ttk.Label(remove_dialog, text="Choose materials to take/remove from the store", font=('Arial', 16))
        heading_label.pack(pady=20)

        #finding the max window dimensions
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        # Set the main window size to fill the screen
        remove_dialog.geometry(f"{screen_width//2}x{screen_height//2}")

        # Create a frame to hold label and Combobox for material
        frame = tk.Frame(remove_dialog)
        frame.pack(padx=10, pady=10, anchor=tk.W)

        # Frame for Material Description
        frame_material_desc = tk.Frame(frame)
        frame_material_desc.pack(anchor=tk.W)

        # label and ComboBox for material selection
        self.material_label = ttk.Label(frame_material_desc, text="Material Description: ", font=('Arial', 10), width=18, anchor=tk.W)
        self.material_label.pack(side=tk.LEFT, padx=10, pady=10)
        self.combodata = ttk.Combobox(frame_material_desc, values=[item[0] for item in self.elements], width=50)
        self.combodata.pack(side=tk.LEFT, padx=10, pady=10)
        self.combodata.bind('<KeyRelease>', self.search_selected_material)
        #calling the tooltip/button description
        materialtooltip_text = "Select materials to remove"
        ToolTip(self.combodata, materialtooltip_text)

        # Frame for Quantity
        frame_quantity = tk.Frame(frame)
        frame_quantity.pack(anchor=tk.W)

        # Entry for quantity input
        self.quantity_label = ttk.Label(frame_quantity, text="Quantity: ", font=('Arial', 10), width=18, anchor=tk.W)
        self.quantity_label.pack(side=tk.LEFT, padx=10, pady=10)
        entry_quantity = ttk.Entry(frame_quantity, width = 15)
        entry_quantity.pack(padx=10, pady=10)
        #calling the tooltip/button description
        qtytooltip_text = "Enter integer quantity"
        ToolTip(entry_quantity, qtytooltip_text)

        frame_entries = tk.Frame(remove_dialog)
        frame_entries.pack(padx=10, pady=10, anchor=tk.W)

        self.used_for_label = ttk.Label(frame_entries, text="Used For: ", font=('Arial', 10), width=12, anchor=tk.E)
        self.used_for_label.grid(row=0, column=0, padx=10, pady=5)
        self.entry_used_for = ttk.Entry(frame_entries, width=30)
        self.entry_used_for.grid(row=0, column=1, padx=10, pady=5)

        self.person_name_2_label = ttk.Label(frame_entries, text="Person Name 2: ", font=('Arial', 10), width=12, anchor=tk.E)
        self.person_name_2_label.grid(row=2, column=0, padx=10, pady=5)
        self.entry_person_name_2 = ttk.Entry(frame_entries, width=40)
        self.entry_person_name_2.grid(row=2, column=1, padx=10, pady=5)

        # Button to confirm removing materials
        confirm_button = ttk.Button(remove_dialog, text="Confirm", command=lambda: self.handle_action("remove", entry_quantity.get()))
        confirm_button.pack(pady=20, anchor=tk.CENTER)   

    def display_material_status(self):
        # Create a new Tkinter window for displaying the table
        table_window = tk.Toplevel(self.root)
        table_window.title('Store Material Status')

        # Read Excel file and create DataFrame
        excel_data = pd.read_excel(self.data_file_path)  # Replace 'your_excel_file.xlsx' with your file name
        self.df = pd.DataFrame(excel_data)

        # Create a frame to hold the search bar
        search_frame = tk.Frame(table_window)
        search_frame.pack(padx=10, pady=10, fill='x')

        # Create and style the search label
        search_label = ttk.Label(search_frame, text="Search Material:", font=('Arial', 10))
        search_label.pack(side=tk.LEFT, padx=5)

        # Create a Combobox for material selection with detailed descriptions
        material_data = [item[0] for item in self.elements]  # Assuming the description is the first element in each item
        self.material_combobox = ttk.Combobox(search_frame, values=material_data, width=70)
        self.material_combobox.pack(side=tk.LEFT, padx=5)
        self.material_combobox.bind('<Return>', self.filter_materials)

        # Create a search button
        search_button = ttk.Button(search_frame, text="Search", command=self.filter_materials)
        search_button.pack(side=tk.LEFT, padx=5)

        # Create Treeview widget
        self.tree = ttk.Treeview(table_window, columns=list(self.df.columns), show='headings')

        # Create Scrollbar
        scrollbar = tk.Scrollbar(table_window, orient='vertical', command=self.tree.yview)
        scrollbar.pack(side='right', fill='y')

        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(expand=True, fill='both')

        # Set Treeview column headings
        for col in self.df.columns:
            self.tree.heading(col, text=col)

        # Set width for the first column (change '0' to the actual column index if needed)
        self.tree.column(self.df.columns[0], width=400)  # Change width as needed

        # Insert data into the Treeview
        for i, row in self.df.iterrows():
            self.tree.insert('', 'end', values=list(row))

    def filter_materials(self, event=None):
        selected_material = self.material_combobox.get().split('(Code:', 1)[0].strip().lower()

        if selected_material == '':
            return

        # Filter the DataFrame based on the selected material
        filtered_df = self.df[self.df['Material Description'].str.lower().str.contains(selected_material)]
        self.update_table(filtered_df)

    def update_table(self, data):
        # Clear existing Treeview data
        self.tree.delete(*self.tree.get_children())

        # Insert updated data into the Treeview
        for i, row in data.iterrows():
            self.tree.insert('', 'end', values=list(row))

    def display_logs(self):
        # Read Excel file and create DataFrame
        if self.log_file_path and self.log_file_path.endswith(('.xlsx', '.xls')):
            excel_data = pd.read_excel(self.log_file_path)
            self.df = pd.DataFrame(excel_data)

            # Create a new Tkinter window for displaying the table
            table_window = tk.Toplevel(self.root)
            table_window.title('Requirements Logs')

            # Create Treeview widget
            self.tree = ttk.Treeview(table_window, columns=['Timestamp', 'Material', 'Material Code', 'Quantity', 'Action', 'Location', 'Person Name'], show='headings')

            # Create Scrollbar
            scrollbar = tk.Scrollbar(table_window, orient='vertical', command=self.tree.yview)
            scrollbar.pack(side='right', fill='y')

            self.tree.configure(yscrollcommand=scrollbar.set)
            self.tree.pack(expand=True, fill='both')

            # Set Treeview column headings
            self.tree.heading('Timestamp', text='Timestamp')
            self.tree.heading('Material', text='Material')
            self.tree.heading('Material Code', text='Material Code')
            self.tree.heading('Quantity', text='Quantity')
            self.tree.heading('Action', text='Action')
            self.tree.heading('Location', text='Location')
            self.tree.heading('Person Name', text='Person Name')

            # Insert data into the Treeview
            for i, row in self.df.iterrows():
                timestamp = row['Timestamp']
                material = row['Material']
                material_code = int(row['Material Code'])
                quantity = int(row['Quantity'])
                action = row['Action']
                location= row['Location']
                person_name = row['Person Name']


                self.tree.insert('','end', values=[timestamp, material, material_code, quantity, action, location, person_name])
        else:
            print("Invalid log file format or no file selected.")

    def handle_action(self, action, quantity):
        selected_material = self.combodata.get()
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if action == "add":
            material_data = [item for item in self.elements if item[0] == selected_material][0]
            material_code = material_data[1]
            current_quantity = int(material_data[3])  # Convert to integer
            new_quantity = current_quantity + int(quantity)

            added_in = self.entry_added_in.get()
            person_name = self.entry_person_name.get()

            log_update = {
                "Timestamp": current_time,
                "Material": selected_material,
                "Material Code": material_code,
                "Quantity": quantity,
                "Action": "Added",
                "Location": added_in,
                "Person Name": person_name
            }

            if hasattr(self, 'wb'):
                # Update the data file for addition
                main_sheet = self.wb.active
                for row in range(2, main_sheet.max_row + 1):
                    material = main_sheet.cell(row=row, column=1).value
                    if material == selected_material:
                        main_sheet.cell(row=row, column=4, value=int(new_quantity))
                        self.wb.save(self.data_file_path)
                        break

            self.load_data_file()
            self.read_log_file()
            self.display_material_status()
            messagebox.showinfo("Data Submitted", f"Material: {selected_material}\nQuantity: {quantity}\nhas been added from store.")

        elif action == "remove":
            material_data = [item for item in self.elements if item[0] == selected_material][0]
            material_code = material_data[1]
            current_quantity = int(material_data[3])
            new_quantity = current_quantity - int(quantity)

            used_for = self.entry_used_for.get()
            person_name_2 = self.entry_person_name_2.get() if self.entry_person_name_2 else ''

            log_update = {
                "Timestamp": current_time,
                "Material": selected_material,
                "Material Code": material_code,
                "Quantity": quantity,
                "Action": "Removed",
                "Location": used_for,
                "Person Name": person_name_2
            }
            if hasattr(self, 'log_data'):
                self.log_data = pd.concat([self.log_data, pd.DataFrame([log_update])], ignore_index=True)
                self.log_data.to_excel(self.log_file_path, index=False)

            if hasattr(self, 'wb'):
                # Update the data file for removal
                main_sheet = self.wb.active
                for row in range(2, main_sheet.max_row + 1):
                    material = main_sheet.cell(row=row, column=1).value
                    if material == selected_material:
                        main_sheet.cell(row=row, column=4, value=int(new_quantity))
                        self.wb.save(self.data_file_path)
                        break
                
            self.load_data_file()
            self.read_log_file()
            self.display_material_status()
            messagebox.showinfo("Data Submitted", f"Material: {selected_material}\nQuantity: {quantity}\nhas been removed from store.")


# Create the Tkinter window and run the app
if __name__ == "__main__":
    root = tk.Tk()
    start_dialog = StartDialog(root)
    root.mainloop()