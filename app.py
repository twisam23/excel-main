import tkinter as tk
from tkinter import ttk,filedialog
from tkinter import messagebox
from openpyxl import load_workbook
import os
from datetime import datetime
import pandas as pd
import json

class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip = None
        self.widget.bind("<Enter>", self.on_enter)
        self.widget.bind("<Leave>", self.on_leave)

    def show_tooltip(self):
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + self.widget.winfo_width() // 2
        y += self.widget.winfo_rooty() + self.widget.winfo_height() + 5

        self.tooltip = tk.Toplevel(self.widget)
        self.tooltip.wm_overrideredirect(True)
        self.tooltip.wm_geometry(f"+{x}+{y}")

        label = tk.Label(self.tooltip, text=self.text, background="#ffffe0", relief="solid", borderwidth=1)
        label.pack(ipadx=2)

    def on_enter(self, event=None):
        self.show_tooltip()

    def on_leave(self, event=None):
        if self.tooltip:
            self.tooltip.destroy()
            self.tooltip = None

class StoreApp:
    def __init__(self, root):
        self.root = root
        self.root.title('Store Supplies')
        
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        self.root.geometry(f"{screen_width}x{screen_height}")

        self.data_file_path = None  # Initializing the data file path
        self.log_file_path = None  # Initializing the log file path

        self.load_stored_paths()

    def load_stored_paths(self):
        try:
            with open('stored_paths.json', 'r') as file:
                stored_paths = json.load(file)
                data_file_path = stored_paths.get("data_file_path")
                log_file_path = stored_paths.get("log_file_path")

            if data_file_path and log_file_path:
                # Check if the stored paths point to existing files
                if os.path.isfile(data_file_path) and os.path.isfile(log_file_path):
                    self.data_file_path = data_file_path
                    self.log_file_path = log_file_path
                    self.show_bf1_dialogue()
                    self.load_data_file()  # Initialize self.wb
                    self.read_log_file()  # Initialize self.log_data
                    return  # Added to exit function after showing BF-1 dialogue
                else:
                    print("Stored paths do not point to existing files.")

            # If paths are missing or invalid, show the initial dialogue
            if not (self.data_file_path and self.log_file_path):
                self.create_initial_dialogue()

        except FileNotFoundError:
            # If the stored_paths.json file is not found, show the initial dialogue
            self.create_initial_dialogue()

    def create_initial_dialogue(self):
        select_files_button = tk.Button(
            self.root,
            text="Select Files",
            command=self.select_files,
            font=('Arial', 16),
            bg='blue',
            fg='white',
            padx=20,
            pady=10,
            relief=tk.RAISED,
            borderwidth=5
        )
        select_files_button.place(relx=0.5, rely=0.4, anchor="center")

    def select_files(self):
        data_file = filedialog.askopenfilename(title="Select Data File", filetypes=[("Excel files", "*.xlsx;*.xls")])
        log_file = filedialog.askopenfilename(title="Select Log File", filetypes=[("Excel files", "*.xlsx;*.xls")])

        if data_file and log_file:
            self.update_paths(data_file, log_file)
            self.show_bf1_dialogue()

    def update_paths(self, new_data_path, new_log_path):
        # Update stored paths with the new paths obtained
        self.data_file_path = new_data_path
        self.log_file_path = new_log_path

        # Save the updated paths to the configuration file (stored_paths.json)
        self.store_paths_to_file()
    
    def load_data_file(self):
        if self.data_file_path and self.data_file_path.endswith(('.xlsx', '.xls')):

            self.wb = load_workbook(filename=self.data_file_path)
            self.sheet = self.wb.active  # or specify a particular sheet by name: self.wb['SheetName']
            self.elements = []

            max_row = self.sheet.max_row  # Find the last row with data
            for row in self.sheet.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=4, values_only=True):
                self.elements.append(row)
                print(row)
        else:
            print("Invalid data file format or no file selected.")

    def read_log_file(self):
        if self.log_file_path and self.log_file_path.endswith(('.xlsx', '.xls')):
            self.log_data = pd.read_excel(self.log_file_path)
            # Process log file data as needed
            print(self.log_data)
        else:
            print("Invalid log file format or no file selected.")
            
    def display_logs(self, storingfile_path):
        if storingfile_path and storingfile_path.endswith(('.xlsx', '.xls')):
            excel_data = pd.read_excel(storingfile_path)
            # Process excel_data as needed
            print(excel_data)
        else:
            print("Invalid file format or no file selected.")

    def search_selected_material(self, event=None):
        if self.combodata:
            value2 = self.combodata.get()

            if value2 == '':
                self.combodata['values'] = [item[0] for item in self.elements]
            else:
                data = []
                for item in self.elements:
                    if value2.lower() in item[0].lower():
                        data.append(item[0])

                self.combodata['values'] = data

    def store_paths_to_file(self):
        stored_paths = {
            "data_file_path": self.data_file_path,
            "log_file_path": self.log_file_path
        }
        with open('stored_paths.json', 'w') as file:
            json.dump(stored_paths, file)

    def show_bf1_dialogue(self):
        self.root.withdraw()
        bf1_dialog = tk.Toplevel(self.root)
        bf1_dialog.title("Store Supplies")

        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        bf1_dialog.geometry(f"{screen_width}x{screen_height}")

        heading_label = ttk.Label(bf1_dialog, text="Store Supplies", font=('Arial', 16))
        heading_label.pack(pady=20)

        display_button = tk.Button(
            bf1_dialog,  # Change this line to use bf1_dialog instead of self.root
            text='Blast Furnace-1',
            command=self.show_bf1_options,
            font=('Arial', 16),
            bg='blue',
            fg='white',
            padx=20,
            pady=10,
            relief=tk.RAISED,
            borderwidth=5
        )
        display_button.place(relx=0.5, rely=0.45, anchor="center")


            #calling the tooltip/button description
        bf1tooltip_text = "Click to enter Blast Furnace-1 options"
        ToolTip(display_button, bf1tooltip_text)

            # Initialize combobox
        self.combodata = None

    
    def show_bf1_options(self):
        # Close the current dialogue box
        self.root.withdraw()

        # Create a new dialog for BF-1 options
        bf1_dialog = tk.Toplevel(self.root)
        bf1_dialog.title("Blast Furnace-1 Options")

        # Create and style the heading
        heading_label = ttk.Label(bf1_dialog, text="Choose Blast Furnace-1 Option", font=('Arial', 16))
        heading_label.pack(pady=20)

        #finding the max window dimensions
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        # Set the main window size to fill the screen
        bf1_dialog.geometry(f"{screen_width}x{screen_height}")

        self.load_data_file()

        # Create a button to Add materials
        add_button = tk.Button(
            bf1_dialog,
            text='Add Materials',
            command=self.add_materials_dialog,
            font=('Arial', 13),  # Change font and size
            fg='black',  # Foreground (text) color
            padx=20,  # Horizontal padding
            pady=10,  # Vertical padding
            relief=tk.RAISED,  # Border style
            borderwidth=5  # Border width
        )
        add_button.place(relx=0.5, rely=0.2, anchor="center")

        #calling the tooltip/button description
        addtooltip_text = "Click to add new Materials to the store"
        ToolTip(add_button, addtooltip_text)

        # Remove Materials Button
        remove_button = tk.Button(
            bf1_dialog,
            text='Remove Materials',
            command=self.remove_materials_dialog,
            font=('Arial', 13),  # Change font and size
            fg='black',  # Foreground (text) color
            padx=20,  # Horizontal padding
            pady=10,  # Vertical padding
            relief=tk.RAISED,  # Border style
            borderwidth=5  # Border width
        )
        remove_button.place(relx=0.5, rely=0.3, anchor="center")

        #calling the tooltip/button description
        removetooltip_text = "Click to take/remove materials from the store"
        ToolTip(remove_button, removetooltip_text)

        # Material Status Button
        status_button = tk.Button(
            bf1_dialog,
            text='Material Status',
            command=self.display_material_status,
            font=('Arial', 13),  # Change font and size
            fg='black',  # Foreground (text) color
            padx=20,  # Horizontal padding
            pady=10,  # Vertical padding
            relief=tk.RAISED,  # Border style
            borderwidth=5  # Border width
        )
        status_button.place(relx=0.5, rely=0.4, anchor="center")

        #calling the tooltip/button description
        statustooltip_text = "Click to view Store Material status"
        ToolTip(status_button, statustooltip_text)

        # Requirements logs Button
        logs_button = tk.Button(
            bf1_dialog,
            text='Store logs',
            command=self.display_logs,
            font=('Arial', 13),  # Change font and size
            fg='black',  # Foreground (text) color
            padx=20,  # Horizontal padding
            pady=10,  # Vertical padding
            relief=tk.RAISED,  # Border style
            borderwidth=5  # Border width
        )
        logs_button.place(relx=0.5, rely=0.5, anchor="center")

        #calling the tooltip/button description
        logstooltip_text = "Click to view Store logs"
        ToolTip(logs_button, logstooltip_text)


    def add_materials_dialog(self):
        # Create a new dialog for Add Materials
        add_dialog = tk.Toplevel(self.root)
        add_dialog.title("Add Materials")
        
        # Create and style the heading
        heading_label = ttk.Label(add_dialog, text="Choose materials to add to the store", font=('Arial', 16))
        heading_label.pack(pady=20)

        #finding the max window dimensions
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        # Set the main window size to fill the screen
        add_dialog.geometry(f"{screen_width//2}x{screen_height//2}")

        # Create a frame to hold label and Combobox for material
        frame = tk.Frame(add_dialog)
        frame.pack(padx=10, pady=10, anchor=tk.W)

        # Frame for Material Description
        frame_material_desc = tk.Frame(frame)
        frame_material_desc.pack(anchor=tk.W)

        # label and ComboBox for material selection
        self.material_label = ttk.Label(frame_material_desc, text="Material Description: ", font=('Arial', 10), width=18, anchor=tk.W)
        self.material_label.pack(side=tk.LEFT, padx=10, pady=10)
        self.combodata = ttk.Combobox(frame_material_desc, values=[item[0] for item in self.elements], width=50)
        self.combodata.pack(side=tk.LEFT, padx=10, pady=10)
        self.combodata.bind('<KeyRelease>', self.search_selected_material)
        #calling the tooltip/button description
        materialtooltip_text = "Select materials to add"
        ToolTip(self.combodata, materialtooltip_text)

        # Frame for Quantity
        frame_quantity = tk.Frame(frame)
        frame_quantity.pack(anchor=tk.W)

        # Entry for quantity input
        self.quantity_label = ttk.Label(frame_quantity, text="Quantity: ", font=('Arial', 10), width=18, anchor=tk.W)
        self.quantity_label.pack(side=tk.LEFT, padx=10, pady=10)
        entry_quantity = ttk.Entry(frame_quantity, width = 15)
        entry_quantity.pack(padx=10, pady=10)
        #calling the tooltip/button description
        qtytooltip_text = "Enter integer quantity"
        ToolTip(entry_quantity, qtytooltip_text)

        # Button to confirm adding materials
        confirm_button = ttk.Button(add_dialog, text="Confirm", command=lambda: self.handle_action("add", entry_quantity.get()))
        confirm_button.pack(pady=20, anchor=tk.CENTER)


    def remove_materials_dialog(self):
        # Create a new dialog for Remove Materials
        remove_dialog = tk.Toplevel(self.root)
        remove_dialog.title("Remove Materials")

        # Create and style the heading
        heading_label = ttk.Label(remove_dialog, text="Choose materials to take/remove from the store", font=('Arial', 16))
        heading_label.pack(pady=20)

        #finding the max window dimensions
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        # Set the main window size to fill the screen
        remove_dialog.geometry(f"{screen_width//2}x{screen_height//2}")

        # Create a frame to hold label and Combobox for material
        frame = tk.Frame(remove_dialog)
        frame.pack(padx=10, pady=10, anchor=tk.W)

        # Frame for Material Description
        frame_material_desc = tk.Frame(frame)
        frame_material_desc.pack(anchor=tk.W)

        # label and ComboBox for material selection
        self.material_label = ttk.Label(frame_material_desc, text="Material Description: ", font=('Arial', 10), width=18, anchor=tk.W)
        self.material_label.pack(side=tk.LEFT, padx=10, pady=10)
        self.combodata = ttk.Combobox(frame_material_desc, values=[item[0] for item in self.elements], width=50)
        self.combodata.pack(side=tk.LEFT, padx=10, pady=10)
        self.combodata.bind('<KeyRelease>', self.search_selected_material)
        #calling the tooltip/button description
        materialtooltip_text = "Select materials to remove"
        ToolTip(self.combodata, materialtooltip_text)

        # Frame for Quantity
        frame_quantity = tk.Frame(frame)
        frame_quantity.pack(anchor=tk.W)

        # Entry for quantity input
        self.quantity_label = ttk.Label(frame_quantity, text="Quantity: ", font=('Arial', 10), width=18, anchor=tk.W)
        self.quantity_label.pack(side=tk.LEFT, padx=10, pady=10)
        entry_quantity = ttk.Entry(frame_quantity, width = 15)
        entry_quantity.pack(padx=10, pady=10)
        #calling the tooltip/button description
        qtytooltip_text = "Enter integer quantity"
        ToolTip(entry_quantity, qtytooltip_text)

        # Button to confirm removing materials
        confirm_button = ttk.Button(remove_dialog, text="Confirm", command=lambda: self.handle_action("remove", entry_quantity.get()))
        confirm_button.pack(pady=20, anchor=tk.CENTER)   

    def display_material_status(self):
        # Create a new Tkinter window for displaying the table
        table_window = tk.Toplevel(self.root)
        table_window.title('Store Material Status')

        # Read Excel file and create DataFrame
        excel_data = pd.read_excel(self.data_file_path)  # Replace 'your_excel_file.xlsx' with your file name
        self.df = pd.DataFrame(excel_data)

        # Create a frame to hold the search bar
        search_frame = tk.Frame(table_window)
        search_frame.pack(padx=10, pady=10, fill='x')

        # Create and style the search label
        search_label = ttk.Label(search_frame, text="Search Material:", font=('Arial', 10))
        search_label.pack(side=tk.LEFT, padx=5)

        # Create a Combobox for material selection with detailed descriptions
        material_data = [desc for desc, _, _, _ in self.elements]
        self.material_combobox = ttk.Combobox(search_frame, values=material_data, width=70)
        self.material_combobox.pack(side=tk.LEFT, padx=5)
        self.material_combobox.bind('<Return>', self.filter_materials)

        # Create a search button
        search_button = ttk.Button(search_frame, text="Search", command=self.filter_materials)
        search_button.pack(side=tk.LEFT, padx=5)

        # Create Treeview widget
        self.tree = ttk.Treeview(table_window, columns=list(self.df.columns), show='headings')

        # Create Scrollbar
        scrollbar = tk.Scrollbar(table_window, orient='vertical', command=self.tree.yview)
        scrollbar.pack(side='right', fill='y')

        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(expand=True, fill='both')

        # Set Treeview column headings
        for col in self.df.columns:
            self.tree.heading(col, text=col)

        # Set width for the first column (change '0' to the actual column index if needed)
        self.tree.column(self.df.columns[0], width=400)  # Change width as needed

        # Insert data into the Treeview
        for i, row in self.df.iterrows():
            self.tree.insert('', 'end', values=list(row))

    def filter_materials(self, event=None):
        selected_material = self.material_combobox.get().split('(Code:', 1)[0].strip().lower()

        if selected_material == '':
            return

        # Filter the DataFrame based on the selected material
        filtered_df = self.df[self.df['Material Description'].str.lower().str.contains(selected_material)]
        self.update_table(filtered_df)

    def update_table(self, data):
        # Clear existing Treeview data
        self.tree.delete(*self.tree.get_children())

        # Insert updated data into the Treeview
        for i, row in data.iterrows():
            self.tree.insert('', 'end', values=list(row))

    def display_logs(self):
        # Read Excel file and create DataFrame
        if self.log_file_path and self.log_file_path.endswith(('.xlsx', '.xls')):
            excel_data = pd.read_excel(self.log_file_path)
            self.df = pd.DataFrame(excel_data)

            # Create a new Tkinter window for displaying the table
            table_window = tk.Toplevel(self.root)
            table_window.title('Requirements Logs')

            # Create Treeview widget
            self.tree = ttk.Treeview(table_window, columns=list(self.df.columns), show='headings')

            # Create Scrollbar
            scrollbar = tk.Scrollbar(table_window, orient='vertical', command=self.tree.yview)
            scrollbar.pack(side='right', fill='y')

            self.tree.configure(yscrollcommand=scrollbar.set)
            self.tree.pack(expand=True, fill='both')

            # Set Treeview column headings
            for col in self.df.columns:
                self.tree.heading(col, text=col)

            # Set width for the first column (change '1' to the actual column index if needed)
            self.tree.column(self.df.columns[1], width=400)  # Change width as needed

            # Insert data into the Treeview
            for i, row in self.df.iterrows():
                self.tree.insert('', 'end', values=list(row))
        else:
            print("Invalid log file format or no file selected.")

    def handle_action(self, action, quantity):
        selected_material = self.combodata.get()
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if action == "add":
            # Find the material data
            material_data = [item for item in self.elements if item[0] == selected_material][0]
            material_data = list(material_data)  # Convert tuple to list for modification
            material_code = material_data[1]
            current_quantity = material_data[3]
            new_quantity = current_quantity + int(quantity)

            # Update the log file for removal
            log_update = [current_time, selected_material, material_code, quantity, "Added"]
            if hasattr(self, 'log_data'):
                self.log_data = self.log_data.append(pd.Series(log_update, index=self.log_data.columns), ignore_index=True)
                self.log_data.to_excel(self.log_file_path, index=False)

            # Update the data file for removal
            if hasattr(self, 'wb'):
                main_sheet = self.wb.active
                for row in range(2, main_sheet.max_row + 1):
                    material = main_sheet.cell(row=row, column=1).value
                    if material == selected_material:
                        main_sheet.cell(row=row, column=4, value=new_quantity)
                        print(f"Material Code: {material_code}, Updated Quantity: {new_quantity}")
                        self.wb.save(self.data_file_path)
                        break
            self.load_data_file()
            self.read_log_file()

            messagebox.showinfo("Data Submitted", f"Material: {selected_material}\nQuantity: {quantity}\nhas been added from store.")

        elif action == "remove":
            # Find the material data
            material_data = [item for item in self.elements if item[0] == selected_material][0]
            material_data = list(material_data)  # Convert tuple to list for modification
            material_code = material_data[1]
            current_quantity = material_data[3]
            new_quantity = current_quantity - int(quantity)

            # Update the log file for removal
            log_update = [current_time, selected_material, material_code, quantity, "Removed"]
            if hasattr(self, 'log_data'):
                self.log_data = self.log_data.append(pd.Series(log_update, index=self.log_data.columns), ignore_index=True)
                self.log_data.to_excel(self.log_file_path, index=False)

            # Update the data file for removal
            if hasattr(self, 'wb'):
                main_sheet = self.wb.active
                for row in range(2, main_sheet.max_row + 1):
                    material = main_sheet.cell(row=row, column=1).value
                    if material == selected_material:
                        main_sheet.cell(row=row, column=4, value=new_quantity)
                        print(f"Material Code: {material_code}, Updated Quantity: {new_quantity}")
                        self.wb.save(self.data_file_path)
                        break
            
            self.load_data_file()
            self.read_log_file()
            messagebox.showinfo("Data Submitted", f"Material: {selected_material}\nQuantity: {quantity}\nhas been removed from store.")

# Create the Tkinter window and run the app
root = tk.Tk()
app = StoreApp(root)
root.mainloop()