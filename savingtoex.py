import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook
from tkinter import *
import tkinter.ttk as ttk
import os

# Load the xlsx file, then store the value of each column in the "elements" list
file_path = r"C:/Users/vyoma/Downloads/excelapp-main (1)/excelapp-main/testdata.xlsx"

if os.path.exists(file_path):
    wb = load_workbook(filename=file_path)
    ws = wb['Sheet1']
    storingfile_path = r"C:/Users/vyoma/Downloads/excelapp-main (1)/excelapp-main\storingfile.xlsx"
    
    try:
        wBook = load_workbook(storingfile_path)
    except FileNotFoundError:
        wBook = load_workbook()
        wBook.save(storingfile_path)

    sheet = wBook.active

    m_row = 1
    m_col = ws.max_column
    MaterialDescription = ws['A2':'A10']
    elements = []

    # to get the list of column values
    for cell in MaterialDescription:
        for x in cell:
            y = x.value
            elements.append(y)
            print(y)

    # search function
    def show_material_info(selected_material):
        # Create a new dialog box for displaying material information
        info_dialog = Toplevel(win)
        info_dialog.title("Material Information")

        # Display material code, base unit of measure, and store quantity
        material_code = base_unit = store_quantity = ""  # Initialize variables
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[0] == selected_material:
                material_code = row[1]
                base_unit = row[2]
                store_quantity = row[3]
                break

        # Label to display information
        ttk.Label(info_dialog, text=f"Material Code: {material_code}").pack(pady=5)
        ttk.Label(info_dialog, text=f"Base Unit: {base_unit}").pack(pady=5)
        ttk.Label(info_dialog, text=f"Store Quantity: {store_quantity}").pack(pady=5)

    def check_input(event):
        if event.keysym == 'Return':
            value2 = event.widget.get()

            if value2 == '':
                combodata['values'] = elements
            else:
                data = []
                for item in elements:
                    if value2.lower() in item.lower():
                        data.append(item)

                combodata['values'] = data

                print("selected value=", combodata.get())
                selected_material = combodata.get()

                # Append the selected material information to the storing file
                data1 = [selected_material]
                sheet.append(data1)
                wBook.save(storingfile_path)

                # Show material information in a new dialog box
                show_material_info(selected_material)

    # Tkinter stuff
    win = Tk()
    clicked = StringVar()

    # label and combobox, binding
    ttk.Label(text="Material Description:").grid(row=1, column=0, padx=10, pady=10)
    combodata = ttk.Combobox(win, values=elements)
    combodata.grid(row=1, column=1, padx=10, pady=10)
    combodata['values'] = elements
    combodata.bind('<KeyRelease>', check_input)
    wBook.save(storingfile_path)

    win.mainloop()
else:
    print("File not found at the specified path.")
